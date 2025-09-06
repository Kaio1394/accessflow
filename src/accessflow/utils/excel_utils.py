from openpyxl import load_workbook
import os
import pandas as pd
from accessflow.core.types import DataTable
from accessflow.core.types import dataframe_to_datatable
from accessflow.exceptions import WorksheetNullObject, WorkbookNullObject, TabNameNotFound

class ExcelUtils:
    def __init__(self, path_excel: str):
        if not path_excel or not path_excel.strip():
           raise ValueError('Invalid value: string <path_excel> is empty.')    
        if not os.path.isfile(path_excel):
            raise FileNotFoundError(f'The Excel [{path_excel}] File not found.')
        self.path_excel = path_excel
        self.workbook = None
        self.worksheet = None
    
    def open(self):
        self.workbook = load_workbook(self.path_excel)  
        
    def set_worksheet(self, tab_name: str):
        if not tab_name:
            raise ValueError("Param 'tab_name' can't be empty.")
        if self.workbook is None:
            raise WorkbookNullObject()
        if tab_name not in self.get_list_worksheets():
            raise TabNameNotFound()       
        self.worksheet = self.workbook[tab_name]
        
    def get_list_worksheets(self) -> list:
        if self.workbook is None:
           raise WorkbookNullObject() 
        return self.workbook.sheetnames
    
    def read_workworksheet(self, tab_name: str) -> DataTable:
        df = pd.read_excel(self.path_excel, sheet_name=tab_name)
        return dataframe_to_datatable(df)

    def close(self):
        if self.workbook is not None:
            self.workbook.save(self.path_excel)
        else:
            raise WorkbookNullObject()